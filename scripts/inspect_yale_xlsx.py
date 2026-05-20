"""
inspect_yale_xlsx.py — Discovery script for Yale Budget Lab's tariff rate
tracker XLSX export. Dumps sheet structure so we can plan integration.

Yale published a downloadable snapshot at:
  https://budgetlab.yale.edu/sites/default/files/2026-04/
    TBL-Data-Tariff-Rate-Tracker-20260401-1.xlsx
(dated 2026-04-01; URL pattern suggests dated successors may follow)

Email from John at TBL (2026-05-17) confirms they're considering a
procedure for ongoing time-series updates. For now we work with this
one-time snapshot.

Usage:
  1. Download the XLSX (above URL) to the project root as
     yale-tariff-data-20260401.xlsx
  2. Activate venv: venv\\Scripts\\activate
  3. Install reader if not already installed:
        pip install openpyxl
  4. Run from project root:
        python -m scripts.inspect_yale_xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("  Run: pip install openpyxl")
    sys.exit(1)


DEFAULT_XLSX = Path("yale-tariff-data-20260401.xlsx")
SAMPLE_ROWS = 5

# Common candidate filenames + locations to auto-discover the download.
DEFAULT_FILENAMES = [
    "yale-tariff-data-20260401.xlsx",
    "TBL-Data-Tariff-Rate-Tracker-20260401-1.xlsx",
]
COMMON_DIRS = [
    Path.cwd(),
    Path.home() / "Downloads",
    Path.home() / "Desktop",
    Path.home() / "Documents",
]


def find_xlsx(explicit: str | None) -> Path | None:
    if explicit:
        p = Path(explicit)
        return p if p.exists() else None
    for d in COMMON_DIRS:
        for fname in DEFAULT_FILENAMES:
            candidate = d / fname
            if candidate.exists():
                return candidate
    return None


def main() -> int:
    explicit = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = find_xlsx(explicit)
    if xlsx_path is None:
        print("ERROR: Cannot find the Yale XLSX in any common location.")
        print("Tried filenames:")
        for f in DEFAULT_FILENAMES:
            print(f"  - {f}")
        print("In directories:")
        for d in COMMON_DIRS:
            print(f"  - {d}")
        print("\nFix one of:")
        print("  (a) Save the XLSX to one of the dirs above with one of the filenames above, or")
        print("  (b) Pass the path explicitly:")
        print("        python -m scripts.inspect_yale_xlsx \"C:\\path\\to\\file.xlsx\"")
        return 1

    print(f"Opening {xlsx_path} (size: {xlsx_path.stat().st_size:,} bytes)\n")
    # data_only=True so we get values, not formulas; read_only is faster for big files.
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)

    print(f"=== Workbook contains {len(wb.sheetnames)} sheet(s) ===\n")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== Sheet: {sheet_name!r} ===")
        # In read_only mode, max_row / max_col aren't always reliable
        # (some files report None). Walk a bounded sample instead.
        rows_iter = ws.iter_rows(values_only=True)

        # Pull the header row (first non-empty row)
        header = None
        for row in rows_iter:
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header = row
                break
        if header is None:
            print("  (empty sheet)\n")
            continue

        print(f"  Columns ({len([c for c in header if c is not None])}):")
        for i, col in enumerate(header):
            print(f"    [{i:2d}] {col!r}")
        print()

        # Sample N rows after the header + count remaining
        sample_rows = []
        count = 0
        for row in rows_iter:
            count += 1
            if len(sample_rows) < SAMPLE_ROWS:
                sample_rows.append(row)

        print(f"  First {len(sample_rows)} data row(s):")
        for r in sample_rows:
            # Truncate strings for readability
            display = tuple(
                (str(v)[:60] + "…") if isinstance(v, str) and len(str(v)) > 60 else v
                for v in r
            )
            print(f"    {display}")
        print()
        print(f"  Total non-header rows in sheet: {count:,}\n")

    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
