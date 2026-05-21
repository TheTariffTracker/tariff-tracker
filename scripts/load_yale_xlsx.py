"""
load_yale_xlsx.py — Parse Yale Budget Lab's Tariff Rate Tracker XLSX snapshot
and upsert all six sheets into Supabase.

Source XLSX (download manually):
  https://budgetlab.yale.edu/sites/default/files/2026-04/
    TBL-Data-Tariff-Rate-Tracker-20260401-1.xlsx

Schema (created by DDL run 2026-05-19):
  yale_etr_overall          — sheet F1, one row per date
  yale_etr_by_authority     — sheet F2, one row per date (wide)
  yale_etr_by_country_group — sheets F3a + F3b combined (long)
  yale_etr_by_sector        — sheet F4 (long)
  yale_etr_sensitivity      — sheet F5, one row per date (wide)
  yale_policy_events        — sheet Policy

All sheets share the same preamble structure (confirmed by inspect_yale_xlsx.py):
  Row 1: figure title (e.g., "Figure 1. Daily Effective Tariff Rate")
  Row 2: Subtitle line
  Row 3: Notes line
  Row 4: Source line
  Row 5: blank
  Row 6: column headers (e.g., "Date", "Revision", "Weighted ETR (%)", ...)
  Row 7+: data

The loader is idempotent — repeated runs upsert on PK. Re-run whenever
Yale publishes a new snapshot.

Usage:
  python -m scripts.load_yale_xlsx <path-to-xlsx>
  python -m scripts.load_yale_xlsx --dry-run <path-to-xlsx>
"""

import sys
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

from scripts.db import upsert


PREAMBLE_ROWS = 5  # title + subtitle + notes + source + blank, BEFORE header row


# ===================== Helpers =====================

def to_date(v) -> str | None:
    """Convert an openpyxl cell value to YYYY-MM-DD string."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # Some sheets may emit ISO strings; pass through if valid-looking.
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        # Last resort: try common US format
        try:
            return datetime.strptime(s, "%m/%d/%Y").date().isoformat()
        except ValueError:
            return None


def to_num(v) -> float | None:
    """Convert a cell value to float, returning None for blanks/non-numeric."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("na", "n/a", "null"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_bool_yesno(v) -> bool | None:
    """Convert "Yes"/"No" / "True"/"False" / 1/0 / TRUE/FALSE to bool."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


def read_sheet_data(ws) -> tuple[list[str], list[tuple]]:
    """Skip preamble rows, return (headers, data_rows)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < PREAMBLE_ROWS + 2:
        return [], []
    headers = list(rows[PREAMBLE_ROWS])  # row index = PREAMBLE_ROWS is header row
    data_rows = rows[PREAMBLE_ROWS + 1 :]
    # Drop trailing fully-empty rows
    while data_rows and all(c is None or c == "" for c in data_rows[-1]):
        data_rows.pop()
    return headers, data_rows


# ===================== Per-sheet parsers =====================

def parse_f1_overall(headers: list, rows: list[tuple]) -> list[dict]:
    """F1: Date, Revision, Weighted ETR (%), Weighted Additional ETR (%),
    Matched Imports ($B), Total Imports ($B)."""
    out = []
    for r in rows:
        d = to_date(r[0])
        if d is None:
            continue
        out.append({
            "date": d,
            "revision": str(r[1]).strip() if r[1] is not None else None,
            "weighted_etr_pct": to_num(r[2]),
            "weighted_additional_etr_pct": to_num(r[3]),
            "matched_imports_bn": to_num(r[4]) if len(r) > 4 else None,
            "total_imports_bn": to_num(r[5]) if len(r) > 5 else None,
        })
    return out


def parse_f2_authority(headers: list, rows: list[tuple]) -> list[dict]:
    """F2: Date, Section 232 (%), Section 301 (%), IEEPA Reciprocal (%),
    IEEPA Fentanyl (%), Section 122 (%), Base Rate (%)."""
    out = []
    for r in rows:
        d = to_date(r[0])
        if d is None:
            continue
        out.append({
            "date": d,
            "section_232_pct": to_num(r[1]),
            "section_301_pct": to_num(r[2]),
            "ieepa_reciprocal_pct": to_num(r[3]),
            "ieepa_fentanyl_pct": to_num(r[4]),
            "section_122_pct": to_num(r[5]),
            "base_rate_pct": to_num(r[6]) if len(r) > 6 else None,
        })
    return out


def parse_country_group(headers: list, rows: list[tuple], group_field_name: str) -> list[dict]:
    """F3a (Date, Group, ETR) and F3b (Date, Partner, ETR) — both feed into
    yale_etr_by_country_group with the dimension column unified as
    `country_group`."""
    out = []
    for r in rows:
        d = to_date(r[0])
        if d is None or r[1] is None:
            continue
        out.append({
            "date": d,
            "country_group": str(r[1]).strip(),
            "etr_pct": to_num(r[2]),
        })
    return out


def parse_f4_sector(headers: list, rows: list[tuple]) -> list[dict]:
    """F4: Date, Sector, ETR (%)."""
    out = []
    for r in rows:
        d = to_date(r[0])
        if d is None or r[1] is None:
            continue
        out.append({
            "date": d,
            "sector": str(r[1]).strip(),
            "etr_pct": to_num(r[2]),
        })
    return out


def parse_f5_sensitivity(headers: list, rows: list[tuple]) -> list[dict]:
    """F5: Date, Main Model (%), 100% Metal Content (%), USMCA 2024 Shares (%)."""
    out = []
    for r in rows:
        d = to_date(r[0])
        if d is None:
            continue
        out.append({
            "date": d,
            "main_model_pct": to_num(r[1]),
            "metal_100pct_pct": to_num(r[2]),
            "usmca_2024_shares_pct": to_num(r[3]) if len(r) > 3 else None,
        })
    return out


def parse_policy(headers: list, rows: list[tuple]) -> list[dict]:
    """Policy: Revision, Effective Date, Policy Event, Major."""
    out = []
    for r in rows:
        rev = str(r[0]).strip() if r[0] is not None else None
        eff = to_date(r[1])
        event = str(r[2]).strip() if len(r) > 2 and r[2] is not None else None
        major = to_bool_yesno(r[3]) if len(r) > 3 else None
        if rev is None or eff is None:
            continue
        out.append({
            "revision": rev,
            "effective_date": eff,
            "policy_event": event,
            "is_major": major,
        })
    return out


# ===================== Main =====================

def main(argv: list[str]) -> int:
    dry_run = "--dry-run" in argv
    args = [a for a in argv if a != "--dry-run"]
    if len(args) < 1:
        print("Usage: python -m scripts.load_yale_xlsx [--dry-run] <path-to-xlsx>")
        return 1
    xlsx_path = Path(args[0])
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        return 1

    if dry_run:
        print("*** DRY RUN — no writes will be made ***")
    print(f"Opening {xlsx_path}")
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)

    plan = [
        # (sheet_name, parser, table_name, on_conflict_cols, friendly_label)
        ("F1",     lambda h, r: parse_f1_overall(h, r),               "yale_etr_overall",          "date",                          "F1 overall ETR"),
        ("F2",     lambda h, r: parse_f2_authority(h, r),             "yale_etr_by_authority",     "date",                          "F2 ETR by authority"),
        ("F3a",    lambda h, r: parse_country_group(h, r, "Group"),   "yale_etr_by_country_group", "date,country_group",            "F3a China vs other"),
        ("F3b",    lambda h, r: parse_country_group(h, r, "Partner"), "yale_etr_by_country_group", "date,country_group",            "F3b ETR by partner"),
        ("F4",     lambda h, r: parse_f4_sector(h, r),                "yale_etr_by_sector",        "date,sector",                   "F4 ETR by sector"),
        ("F5",     lambda h, r: parse_f5_sensitivity(h, r),           "yale_etr_sensitivity",      "date",                          "F5 sensitivity"),
        ("Policy", lambda h, r: parse_policy(h, r),                   "yale_policy_events",        "revision,effective_date",       "Policy events"),
    ]

    total_inserted = 0
    for sheet_name, parser, table, on_conflict, label in plan:
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP {label}: sheet {sheet_name!r} not in workbook")
            continue
        ws = wb[sheet_name]
        headers, data_rows = read_sheet_data(ws)
        mapped = parser(headers, data_rows)
        print(f"  {label}: {len(data_rows)} data rows → {len(mapped)} mapped")
        if not mapped:
            continue
        if dry_run:
            print(f"    DRY RUN: would upsert {len(mapped)} rows to {table}.")
            # show a sample
            print(f"    Sample first row: {mapped[0]}")
            if len(mapped) > 1:
                print(f"    Sample last row:  {mapped[-1]}")
        else:
            upsert(
                table,
                mapped,
                on_conflict=on_conflict,
                progress_label=f"  Upserted {table}",
            )
            total_inserted += len(mapped)

    wb.close()
    if dry_run:
        print("\nDONE (dry run, no changes).")
    else:
        print(f"\nDONE. Upserted {total_inserted:,} rows across all Yale tables.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
